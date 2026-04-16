from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional
from urllib.parse import quote

import openpyxl
import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from PIL import Image
    from io import BytesIO
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

BASE_URL = "https://zukan.pokemon.co.jp/"
DETAIL_PATH_RE = re.compile(r"/detail/([0-9]{4}(?:-[0-9]+)?)")
IMAGE_URL_RE = re.compile(r"https?://[^\s'\"<>]+?\.(?:png|jpg|jpeg|webp)(?:\?[^\s'\"<>]*)?", re.IGNORECASE)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
    "Referer": BASE_URL,
}

FORM_HINTS = {
    "Alola": ["アローラ", "アローラのすがた"],
    "Galar": ["ガラル", "ガラルのすがた"],
    "Hisui": ["ヒスイ", "ヒスイのすがた"],
    "Paldea": ["パルデア", "パルデアのすがた"],
    "Origin": ["オリジン", "オリジンフォルム"],
    "Normal": ["ノーマル", "ノーマルフォルム", "のすがた"],
    "Plant": ["くさきのミノ"],
    "Sandy": ["すなちのミノ"],
    "Trash": ["ゴミのミノ"],
    "WestSea": ["にしのうみ"],
    "EastSea": ["ひがしのうみ"],
    "Male": ["オス", "♂"],
    "Female": ["メス", "♀"],
    "Red": ["あかすじ"],
    "Blue": ["あおすじ"],
    "White": ["しろすじ", "ホワイト"],
    "Black": ["ブラック"],
    "10Percent": ["10%", "１０％"],
    "50Percent": ["50%", "５０％"],
    "Complete": ["パーフェクト"],
    "Baile": ["めらめら", "バイレ"],
    "Pau": ["ふらふら"],
    "PomPom": ["ぱちぱち"],
    "Sensu": ["まいまい"],
    "Midday": ["まひる"],
    "Dusk": ["たそがれ"],
    "Midnight": ["まよなか"],
    "Meteor": ["りゅうせい", "メテオ"],
    "Core": ["コア"],
    "DawnWings": ["あかつきのつばさ"],
    "DuskMane": ["たそがれのたてがみ"],
    "Ultra": ["ウルトラ"],
    "Amped": ["ハイなすがた", "ハイ"],
    "LowKey": ["ローなすがた", "ロー"],
    "SingleStrike": ["いちげき", "いちげきのかた"],
    "RapidStrike": ["れんげき", "れんげきのかた"],
    "Bloodmoon": ["アカツキ", "あかつき"],
    "Chest": ["はこフォルム", "はこ"],
    "Roaming": ["とほフォルム", "とほ"],
    "Terastal": ["テラスタル"],
}

GENERIC_IMAGE_BLOCKLIST = [
    "logo", "twitter", "facebook", "line", "sns", "icon", "favicon", "apple-touch",
    "loading", "spinner", "banner", "share", "ogp-default", "pagetop", "common",
]

@dataclass
class RowData:
    row_num: int
    row_id: str
    image_path: str
    answer: str
    aliases: list[str]

    @property
    def base_no(self) -> str:
        return self.row_id.split("-", 1)[0]

    @property
    def suffix(self) -> Optional[str]:
        return self.row_id.split("-", 1)[1] if "-" in self.row_id else None


@dataclass
class DetailPage:
    detail_id: str
    url: str
    title: str
    image_url: Optional[str]
    image_source: Optional[str]


def normalize_text(text: str) -> str:
    text = unicodedata.normalize("NFKC", text or "")
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("　", " ")
    text = text.replace("／", "/")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def canonical_spaces(text: str) -> str:
    text = normalize_text(text)
    for ch in ["(", ")", "・", "/", "／", "-", "_", ":", "：", "、", ",", "　"]:
        text = text.replace(ch, " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def compact(text: str) -> str:
    return canonical_spaces(text).replace(" ", "")


def split_name_and_form(answer: str) -> tuple[str, str]:
    s = normalize_text(answer)
    parts = re.split(r"\s+", s, maxsplit=1)
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def build_variants(row: RowData) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []

    def add(value: str) -> None:
        value = normalize_text(value)
        if value and value not in seen:
            seen.add(value)
            out.append(value)

    add(row.answer)
    base_name, form_name = split_name_and_form(row.answer)
    add(base_name)
    if form_name:
        add(f"{base_name} {form_name}")
        add(f"{base_name}（{form_name}）")
        add(f"{base_name}({form_name})")
        add(f"{base_name}{form_name}")
        add(form_name)

    for alias in row.aliases:
        add(alias)

    if row.suffix:
        for hint in FORM_HINTS.get(row.suffix, []):
            add(f"{base_name} {hint}")
            add(f"{base_name}（{hint}）")
            add(f"{base_name}{hint}")
            add(f"{hint}{base_name}")

    return out


def title_from_html(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")

    meta = soup.find("meta", {"property": "og:title"})
    if meta and meta.get("content"):
        title = normalize_text(meta["content"])
        title = re.sub(r"\s*[\|\|｜]\s*ポケモンずかん\s*$", "", title)
        return title

    if soup.title and soup.title.string:
        title = normalize_text(soup.title.string)
        title = re.sub(r"\s*[|｜]\s*ポケモンずかん\s*$", "", title)
        return title

    for tag in soup.find_all(["h1", "h2"]):
        text = normalize_text(tag.get_text(" ", strip=True))
        if text:
            return text

    return ""


def iter_strings(obj):
    if isinstance(obj, dict):
        for value in obj.values():
            yield from iter_strings(value)
    elif isinstance(obj, list):
        for value in obj:
            yield from iter_strings(value)
    elif isinstance(obj, str):
        yield obj


def score_image_url(url: str) -> int:
    u = url.lower()
    score = 0
    if "zukan.pokemon.co.jp" in u:
        score += 40
    if any(token in u for token in ["/img/", "/images/", "/pokemon/", "/zukan/"]):
        score += 10
    if u.endswith(".png") or ".png?" in u:
        score += 8
    if u.endswith(".webp") or ".webp?" in u:
        score += 6
    if u.endswith(".jpg") or u.endswith(".jpeg") or ".jpg?" in u or ".jpeg?" in u:
        score += 4
    if any(token in u for token in ["pokemon", "monster", "detail"]):
        score += 5
    if any(token in u for token in GENERIC_IMAGE_BLOCKLIST):
        score -= 100
    return score


def pick_best_image(urls: Iterable[str]) -> tuple[Optional[str], Optional[str]]:
    cleaned: list[str] = []
    seen: set[str] = set()
    for url in urls:
        url = normalize_text(url)
        if not url or url in seen:
            continue
        seen.add(url)
        cleaned.append(url)

    if not cleaned:
        return None, None

    scored = sorted(((score_image_url(u), u) for u in cleaned), reverse=True)
    best_score, best_url = scored[0]
    if best_score < 0:
        return None, None
    return best_url, "best_guess"


def extract_image_url_from_detail_html(html: str) -> tuple[Optional[str], Optional[str]]:
    soup = BeautifulSoup(html, "html.parser")
    candidates: list[tuple[str, str]] = []

    for attrs, label in [
        ({"property": "og:image"}, "og:image"),
        ({"name": "twitter:image"}, "twitter:image"),
        ({"property": "twitter:image"}, "twitter:image"),
    ]:
        meta = soup.find("meta", attrs)
        if meta and meta.get("content"):
            candidates.append((meta["content"], label))

    for script in soup.find_all("script"):
        script_type = (script.get("type") or "").lower()
        raw = script.string or script.get_text() or ""
        if not raw.strip():
            continue

        if "json" in script_type:
            try:
                data = json.loads(raw)
                for s in iter_strings(data):
                    if IMAGE_URL_RE.search(s):
                        for hit in IMAGE_URL_RE.findall(s):
                            candidates.append((hit, "json"))
            except Exception:
                pass

        for hit in IMAGE_URL_RE.findall(raw):
            candidates.append((hit, "script_regex"))

    for img in soup.find_all("img"):
        for attr in ["src", "data-src", "data-original", "data-lazy-src"]:
            if img.get(attr):
                candidates.append((img[attr], f"img:{attr}"))
        if img.get("srcset"):
            for part in img["srcset"].split(","):
                url = part.strip().split(" ")[0]
                if url:
                    candidates.append((url, "img:srcset"))

    best_url, _ = pick_best_image(url for url, _ in candidates)
    if not best_url:
        return None, None
    for url, source in candidates:
        if normalize_text(url) == best_url:
            return best_url, source
    return best_url, "best_guess"


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)

    retry = Retry(
        total=4,
        connect=4,
        read=4,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "HEAD"],
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def fetch_detail_page(
    session: requests.Session,
    detail_id: str,
    timeout: float,
) -> Optional[DetailPage]:
    url = f"{BASE_URL}detail/{detail_id}"
    res = session.get(url, timeout=timeout, allow_redirects=True)
    if res.status_code != 200:
        return None

    match = DETAIL_PATH_RE.search(res.url)
    if not match:
        return None

    actual_id = match.group(1)
    title = title_from_html(res.text)
    if not title or title.startswith("トップページ"):
        return None

    image_url, image_source = extract_image_url_from_detail_html(res.text)
    return DetailPage(
        detail_id=actual_id,
        url=res.url,
        title=title,
        image_url=image_url,
        image_source=image_source,
    )


def score_detail(row: RowData, page: DetailPage) -> int:
    if page.detail_id.split("-", 1)[0] != row.base_no:
        return -10_000

    score = 0
    title_compact = compact(page.title)
    base_name, form_name = split_name_and_form(row.answer)
    variants = build_variants(row)
    variant_compacts = [compact(v) for v in variants if compact(v)]

    if title_compact in variant_compacts:
        score += 350

    for vc in variant_compacts:
        if not vc:
            continue
        if vc == title_compact:
            score += 250
        elif vc in title_compact:
            score += 90
        elif title_compact in vc:
            score += 70

    base_compact = compact(base_name)
    if base_compact and base_compact in title_compact:
        score += 40

    if row.suffix:
        if "-" in page.detail_id:
            score += 20
        else:
            score += 10 if row.suffix in {"Normal", "Amped"} else -10
    else:
        if "-" not in page.detail_id:
            score += 40
        else:
            score -= 20

    if form_name:
        for token in canonical_spaces(form_name).split():
            token_c = compact(token)
            if token_c and token_c in title_compact:
                score += 35

    if row.suffix:
        for hint in FORM_HINTS.get(row.suffix, []):
            hint_c = compact(hint)
            if hint_c and hint_c in title_compact:
                score += 35

    # Normal / base-form rows often live on the plain base URL.
    if row.suffix == "Normal" and page.detail_id == row.base_no:
        score += 60
    if row.suffix in {"Amped", "Meteor", "Chest"} and page.detail_id == row.base_no:
        score += 45

    return score


def choose_best_page(row: RowData, pages: list[DetailPage]) -> tuple[Optional[DetailPage], list[tuple[int, DetailPage]]]:
    scored = sorted(((score_detail(row, p), p) for p in pages), key=lambda x: x[0], reverse=True)
    if not scored:
        return None, []
    if scored[0][0] < 120:
        return None, scored
    return scored[0][1], scored


def extract_detail_ids_from_html(html: str, base_no: str) -> list[str]:
    found = []
    seen = set()
    for detail_id in DETAIL_PATH_RE.findall(html):
        if detail_id.split("-", 1)[0] != base_no:
            continue
        if detail_id not in seen:
            seen.add(detail_id)
            found.append(detail_id)
    return found


def search_detail_ids(
    session: requests.Session,
    row: RowData,
    timeout: float,
    sleep_sec: float,
) -> list[str]:
    ids: list[str] = []
    seen: set[str] = set()

    for term in build_variants(row):
        url = f"{BASE_URL}?word={quote(term)}&no_min={int(row.base_no)}&no_max={int(row.base_no)}"
        res = session.get(url, timeout=timeout)
        if res.status_code == 200:
            for detail_id in extract_detail_ids_from_html(res.text, row.base_no):
                if detail_id not in seen:
                    seen.add(detail_id)
                    ids.append(detail_id)
        time.sleep(sleep_sec)

    return ids


def read_rows(xlsx_path: Path) -> list[RowData]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    image_col = "image" if "image" in idx else "images" if "images" in idx else None
    required = ["id", "answer", "aliases"]
    missing = [col for col in required if col not in idx]
    if image_col is None:
        missing.append("image/images")
    if missing:
        raise ValueError(f"必須列が不足しています: {', '.join(missing)}")

    rows: list[RowData] = []
    for row_num, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_id = normalize_text(str(values[idx["id"]] or ""))
        image_path = normalize_text(str(values[idx[image_col]] or ""))
        answer = normalize_text(str(values[idx["answer"]] or ""))
        aliases_raw = normalize_text(str(values[idx["aliases"]] or ""))

        if not row_id or not image_path or not answer:
            continue

        aliases = [normalize_text(x) for x in aliases_raw.split("|") if normalize_text(x)]
        rows.append(
            RowData(
                row_num=row_num,
                row_id=row_id,
                image_path=image_path,
                answer=answer,
                aliases=aliases,
            )
        )
    return rows


def save_image_bytes(content: bytes, target_path: Path, source_url: str) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    lower = source_url.lower()

    if lower.endswith(".png") or ".png?" in lower:
        target_path.write_bytes(content)
        return

    if not PIL_AVAILABLE:
        target_path.write_bytes(content)
        return

    with Image.open(BytesIO(content)) as im:
        if im.mode not in ("RGBA", "RGB"):
            im = im.convert("RGBA")
        im.save(target_path, format="PNG")


def zip_images(images_dir: Path, zip_path: Path) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(images_dir.rglob("*")):
            if path.is_file():
                zf.write(path, arcname=path.relative_to(images_dir.parent))


def main() -> int:
    parser = argparse.ArgumentParser(
        description="現在のExcel（id / image or images / answer / aliases）準拠で、ポケモンずかんから画像を取得して ZIP 化します。"
    )
    parser.add_argument("xlsx", help="入力Excelファイル")
    parser.add_argument("--output-dir", default="pokemon_zukan_download_output", help="出力フォルダ")
    parser.add_argument("--zip-name", default="pokemon_zukan_images.zip", help="作成するZIP名")
    parser.add_argument("--timeout", type=float, default=30.0, help="HTTPタイムアウト秒数")
    parser.add_argument("--sleep", type=float, default=0.2, help="リクエスト間の待機秒数")
    parser.add_argument("--max-form-index", type=int, default=12, help="すがた違い候補として試す detail/xxxx-N の最大 N")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"入力ファイルが見つかりません: {xlsx_path}", file=sys.stderr)
        return 1

    output_dir = Path(args.output_dir)
    images_dir = output_dir / "images"
    report_path = output_dir / "download_report.csv"
    zip_path = output_dir.parent / args.zip_name

    output_dir.mkdir(parents=True, exist_ok=True)
    images_dir.mkdir(parents=True, exist_ok=True)

    rows = read_rows(xlsx_path)
    session = make_session()

    page_cache: dict[str, Optional[DetailPage]] = {}
    base_cache: dict[str, list[DetailPage]] = {}

    def get_page(detail_id: str) -> Optional[DetailPage]:
        if detail_id not in page_cache:
            page_cache[detail_id] = fetch_detail_page(session, detail_id, args.timeout)
            time.sleep(args.sleep)
        return page_cache[detail_id]

    def get_pages_for_base(row: RowData) -> list[DetailPage]:
        if row.base_no in base_cache:
            return base_cache[row.base_no]

        pages: list[DetailPage] = []
        seen_ids: set[str] = set()

        # まずは素の図鑑番号。
        for probe_id in [row.base_no] + [f"{row.base_no}-{i}" for i in range(1, args.max_form_index + 1)]:
            page = get_page(probe_id)
            if page is None:
                continue
            if page.detail_id not in seen_ids:
                seen_ids.add(page.detail_id)
                pages.append(page)

        # 念のため検索ページからも detail id を拾う。
        for detail_id in search_detail_ids(session, row, args.timeout, args.sleep):
            page = get_page(detail_id)
            if page and page.detail_id not in seen_ids:
                seen_ids.add(page.detail_id)
                pages.append(page)

        base_cache[row.base_no] = pages
        return pages

    with report_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "row_num",
            "id",
            "answer",
            "selected_detail_id",
            "selected_detail_url",
            "page_title",
            "image_url",
            "image_source",
            "status",
            "note",
        ])

        total = len(rows)
        for idx, row in enumerate(rows, start=1):
            target_path = output_dir / row.image_path
            print(f"[{idx}/{total}] {row.row_id} {row.answer}")

            if target_path.exists():
                writer.writerow([
                    row.row_num, row.row_id, row.answer, "", "", "", "", "", "skipped_existing", "既存ファイルを利用",
                ])
                print("  -> skipped_existing")
                continue

            try:
                pages = get_pages_for_base(row)
                selected, scored = choose_best_page(row, pages)

                if selected is None:
                    top = " | ".join([f"{p.detail_id}:{p.title}:{s}" for s, p in scored[:5]])
                    writer.writerow([
                        row.row_num, row.row_id, row.answer, "", "", "", "", "", "not_found", top,
                    ])
                    print("  -> 候補を確定できませんでした")
                    continue

                image_url = selected.image_url
                image_source = selected.image_source
                if not image_url:
                    # 念のため再取得して HTML から取り直す
                    refreshed = fetch_detail_page(session, selected.detail_id, args.timeout)
                    time.sleep(args.sleep)
                    if refreshed:
                        selected = refreshed
                        image_url = selected.image_url
                        image_source = selected.image_source

                if not image_url:
                    writer.writerow([
                        row.row_num, row.row_id, row.answer,
                        selected.detail_id, selected.url, selected.title, "", "", "detail_parse_failed", "",
                    ])
                    print("  -> 詳細ページから画像URLを取得できませんでした")
                    continue

                image_res = session.get(image_url, timeout=args.timeout)
                image_res.raise_for_status()
                save_image_bytes(image_res.content, target_path, image_url)
                time.sleep(args.sleep)

                writer.writerow([
                    row.row_num, row.row_id, row.answer,
                    selected.detail_id, selected.url, selected.title,
                    image_url, image_source or "", "downloaded", "",
                ])
                print(f"  -> saved: {target_path}")

            except Exception as e:
                writer.writerow([
                    row.row_num, row.row_id, row.answer, "", "", "", "", "", "error", repr(e),
                ])
                print(f"  -> error: {e}")

    zip_images(images_dir, zip_path)
    print(f"完了: {zip_path}")
    print(f"レポート: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

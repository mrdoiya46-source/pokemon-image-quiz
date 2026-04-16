from __future__ import annotations

import argparse
import csv
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

BASE_URL = "https://zukan.pokemon.co.jp"

BLOCKLIST = (
    "logo", "twitter", "facebook", "line", "sns", "icon", "favicon",
    "pagetop", "banner", "share", "header", "footer", "common"
)

DEFAULT_MAX_SUFFIX = 12

# Excelの id サフィックスと、ずかんページの表記の対応
SUFFIX_HINTS = {
    "Alola": ["アローラ", "アローラのすがた"],
    "Galar": ["ガラル", "ガラルのすがた"],
    "Hisui": ["ヒスイ", "ヒスイのすがた"],
    "Paldea": ["パルデア", "パルデアのすがた"],
    "Origin": ["オリジン", "オリジンフォルム"],
    "Attack": ["アタック"],
    "Defense": ["ディフェンス"],
    "Speed": ["スピード"],
    "Plant": ["くさきのミノ"],
    "Sandy": ["すなちのミノ"],
    "Trash": ["ゴミのミノ"],
    "WestSea": ["にしのうみ"],
    "EastSea": ["ひがしのうみ"],
    "Red": ["あかすじ"],
    "Blue": ["あおすじ"],
    "White": ["しろすじ", "ホワイト"],
    "Black": ["ブラック"],
    "Male": ["オス", "♂"],
    "Female": ["メス", "♀"],
    "10Percent": ["10%", "１０％"],
    "50Percent": ["50%", "５０％"],
    "Complete": ["パーフェクト"],
    "Baile": ["めらめら"],
    "Pau": ["ふらふら"],
    "PomPom": ["ぱちぱち"],
    "Sensu": ["まいまい"],
    "Midday": ["まひる"],
    "Midnight": ["まよなか"],
    "Dusk": ["たそがれ"],
    "Meteor": ["りゅうせい", "メテオ"],
    "Core": ["コア"],
    "DawnWings": ["あかつきのつばさ"],
    "DuskMane": ["たそがれのたてがみ"],
    "Ultra": ["ウルトラ"],
    "Amped": ["ハイなすがた", "ハイ"],
    "LowKey": ["ローなすがた", "ロー"],
    "SingleStrike": ["いちげき", "いちげきのかた"],
    "RapidStrike": ["れんげき", "れんげきのかた"],
    "Bloodmoon": ["アカツキ", "あかつき"],
    "Chest": ["はこフォルム", "はこ"],
    "Roaming": ["とほフォルム", "とほ"],
    "Terastal": ["テラスタル"],
    "Stellar": ["ステラ"],
    "Normal": ["ノーマル", "ノーマルフォルム"],
}

@dataclass
class RowData:
    row_num: int
    row_id: str
    image_path: str
    answer: str
    aliases: list[str]

    @property
    def base_no(self) -> str:
        return self.row_id.split("-", 1)[0]

    @property
    def suffix(self) -> Optional[str]:
        return self.row_id.split("-", 1)[1] if "-" in self.row_id else None


@dataclass
class CandidatePage:
    detail_id: str
    url: str
    title: str


def normalize_text(text: str) -> str:
    text = unicodedata.normalize("NFKC", text or "")
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("　", " ")
    text = text.replace("／", "/")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def compact(text: str) -> str:
    text = normalize_text(text).lower()
    for ch in (" ", "(", ")", "・", "/", "／", "-", "_", ":", "：", "、", ",", ".", "．", "’", "'", "　"):
        text = text.replace(ch, "")
    return text


def split_name_and_form(answer: str) -> tuple[str, str]:
    s = normalize_text(answer)
    parts = re.split(r"\s+", s, maxsplit=1)
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def row_name_variants(row: RowData) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []

    def add(v: str) -> None:
        v = normalize_text(v)
        if v and v not in seen:
            seen.add(v)
            out.append(v)

    add(row.answer)
    base_name, form_name = split_name_and_form(row.answer)
    add(base_name)

    if form_name:
        add(f"{base_name} {form_name}")
        add(f"{base_name}（{form_name}）")
        add(f"{base_name}({form_name})")
        add(f"{base_name}{form_name}")
        add(form_name)

    for alias in row.aliases:
        add(alias)

    if row.suffix and row.suffix in SUFFIX_HINTS:
        for hint in SUFFIX_HINTS[row.suffix]:
            add(f"{base_name} {hint}")
            add(f"{base_name}（{hint}）")
            add(f"{base_name}{hint}")
            add(f"{hint}{base_name}")

    return out


def score_title(row: RowData, title: str) -> int:
    title_n = compact(title)
    base_name, form_name = split_name_and_form(row.answer)
    base_n = compact(base_name)

    score = 0

    # 完全一致を最重視
    for v in row_name_variants(row):
        vn = compact(v)
        if not vn:
            continue
        if title_n == vn:
            score += 500
        elif vn in title_n:
            score += 80
        elif title_n in vn:
            score += 40

    if base_n and base_n in title_n:
        score += 50

    if form_name:
        for token in normalize_text(form_name).split():
            tn = compact(token)
            if tn and tn in title_n:
                score += 35

    if row.suffix:
        for hint in SUFFIX_HINTS.get(row.suffix, []):
            hn = compact(hint)
            if hn and hn in title_n:
                score += 40
    else:
        # サフィックスがない行は base URL そのものを優先
        score += 20

    # 余計なフォーム語が入っている場合は減点
    form_tokens = [
        "アローラ", "ガラル", "ヒスイ", "パルデア", "オリジン",
        "アタック", "ディフェンス", "スピード",
        "ノーマル", "ハイ", "ロー", "まひる", "まよなか", "たそがれ",
        "ウルトラ", "テラスタル", "いちげき", "れんげき",
    ]
    if not row.suffix:
        for tok in form_tokens:
            if compact(tok) in title_n:
                score -= 15

    return score


def extract_excel_rows(xlsx_path: Path) -> list[RowData]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers)}

    image_col = "image" if "image" in idx else "images" if "images" in idx else None
    required = ["id", "answer", "aliases"]
    missing = [c for c in required if c not in idx]
    if image_col is None:
        missing.append("image/images")
    if missing:
        raise ValueError(f"必須列が不足しています: {', '.join(missing)}")

    rows: list[RowData] = []
    for row_num, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_id = normalize_text(str(values[idx["id"]] or ""))
        image_path = normalize_text(str(values[idx[image_col]] or ""))
        answer = normalize_text(str(values[idx["answer"]] or ""))
        aliases_raw = normalize_text(str(values[idx["aliases"]] or ""))

        if not row_id or not image_path or not answer:
            continue

        aliases = [normalize_text(x) for x in aliases_raw.split("|") if normalize_text(x)]
        rows.append(RowData(row_num=row_num, row_id=row_id, image_path=image_path, answer=answer, aliases=aliases))
    return rows


def safe_title(text: str) -> str:
    text = normalize_text(text)
    text = re.sub(r"\s*[|｜].*$", "", text)
    return text


def collect_candidate_pages(context, row: RowData, max_suffix: int) -> list[CandidatePage]:
    page = context.new_page()
    page.set_default_timeout(20000)

    candidates: list[CandidatePage] = []
    seen: set[str] = set()

    def add_candidate(url: str) -> None:
        if url in seen:
            return
        seen.add(url)
        try:
            page.goto(url, wait_until="domcontentloaded")
            page.wait_for_timeout(400)
            final_url = page.url
            m = re.search(r"/detail/([0-9]{4}(?:-[0-9]+)?)", final_url)
            if not m:
                return
            detail_id = m.group(1)
            title = safe_title(page.title())
            if not title:
                return
            if detail_id.split("-", 1)[0] != row.base_no:
                return
            candidates.append(CandidatePage(detail_id=detail_id, url=final_url, title=title))
        except Exception:
            return

    # まずは決め打ち
    add_candidate(f"{BASE_URL}/detail/{row.base_no}")
    for i in range(1, max_suffix + 1):
        add_candidate(f"{BASE_URL}/detail/{row.base_no}-{i}")

    # 次に検索で補完
    for term in row_name_variants(row):
        try:
            search_url = f"{BASE_URL}/?word={term}&no_min={int(row.base_no)}&no_max={int(row.base_no)}"
            page.goto(search_url, wait_until="domcontentloaded")
            page.wait_for_timeout(500)
            hrefs = page.eval_on_selector_all(
                'a[href*="/detail/"]',
                """els => els.map(el => el.href).filter(Boolean)"""
            )
            for href in hrefs:
                m = re.search(r"/detail/([0-9]{4}(?:-[0-9]+)?)", href)
                if not m:
                    continue
                detail_id = m.group(1)
                if detail_id.split("-", 1)[0] != row.base_no:
                    continue
                add_candidate(href)
        except Exception:
            continue

    page.close()

    # detail_id 単位で重複排除
    dedup: dict[str, CandidatePage] = {}
    for c in candidates:
        dedup[c.detail_id] = c
    return list(dedup.values())


def choose_candidate(row: RowData, candidates: list[CandidatePage]) -> tuple[Optional[CandidatePage], list[tuple[int, CandidatePage]]]:
    scored = sorted(
        [(score_title(row, c.title), c) for c in candidates],
        key=lambda x: (x[0], x[1].detail_id == row.base_no),
        reverse=True,
    )
    if not scored:
        return None, []

    # 厳しめに判定する
    best_score, best = scored[0]

    # base 行は plain detail/base を優先
    if row.suffix is None:
        for score, c in scored:
            if c.detail_id == row.base_no and compact(split_name_and_form(row.answer)[0]) in compact(c.title):
                return c, scored

    if best_score < 180:
        return None, scored

    # 2位と僅差なら無理に確定しない
    if len(scored) >= 2 and best_score - scored[1][0] < 35 and best_score < 450:
        return None, scored

    return best, scored


JS_PICK_MAIN_VISUAL = r"""
() => {
  const block = /logo|twitter|facebook|line|sns|icon|favicon|pagetop|banner|share|header|footer|common/i;

  const isVisible = (el) => {
    const style = window.getComputedStyle(el);
    if (!style) return false;
    if (style.display === 'none' || style.visibility === 'hidden' || Number(style.opacity || '1') === 0) return false;
    const rect = el.getBoundingClientRect();
    if (rect.width < 120 || rect.height < 120) return false;
    if (rect.bottom <= 0 || rect.right <= 0) return false;
    return true;
  };

  const candidates = [];

  const pushCandidate = (kind, el, src) => {
    if (!el || !isVisible(el)) return;
    const rect = el.getBoundingClientRect();
    const text = [src || '', el.id || '', el.className || '', el.getAttribute('alt') || '', el.getAttribute('aria-label') || ''].join(' ');
    if (block.test(text)) return;

    let score = rect.width * rect.height;
    if (rect.top < window.innerHeight * 0.7) score += 50000;
    if (rect.top < 900) score += 30000;
    if (rect.left < window.innerWidth * 0.8) score += 10000;
    if (rect.width >= 250 && rect.height >= 250) score += 40000;
    if (rect.width >= 350 && rect.height >= 350) score += 50000;
    if (kind === 'img') score += 10000;
    if (src && /pokemon|zukan|detail/i.test(src)) score += 15000;
    if (src && /\.(png|webp|jpg|jpeg)(\?|$)/i.test(src)) score += 8000;

    candidates.push({
      kind,
      src: src || '',
      selectorHint: el.id ? '#' + el.id : (el.className ? '.' + String(el.className).split(/\s+/)[0] : el.tagName.toLowerCase()),
      x: rect.left,
      y: rect.top,
      width: rect.width,
      height: rect.height,
      score,
    });
  };

  document.querySelectorAll('img').forEach(img => {
    pushCandidate('img', img, img.currentSrc || img.src || '');
  });

  document.querySelectorAll('*').forEach(el => {
    const bg = window.getComputedStyle(el).backgroundImage || '';
    if (bg && bg !== 'none') {
      pushCandidate('bg', el, bg);
    }
  });

  candidates.sort((a, b) => b.score - a.score);
  return candidates;
}
"""


def save_main_visual(context, detail_url: str, out_path: Path) -> tuple[bool, str]:
    page = context.new_page()
    page.set_default_timeout(30000)
    try:
        page.goto(detail_url, wait_until="networkidle")
    except PlaywrightTimeoutError:
        page.goto(detail_url, wait_until="domcontentloaded")
        page.wait_for_timeout(1500)

    page.wait_for_timeout(800)

    candidates = page.evaluate(JS_PICK_MAIN_VISUAL)

    # まず img を優先
    best_img = next((c for c in candidates if c["kind"] == "img"), None)

    if best_img:
        src = best_img.get("src", "")
        locator = page.locator(f'img[src="{src}"]')
        if locator.count() > 0:
            locator.first.screenshot(path=str(out_path))
            page.close()
            return True, f'img:{src}'

    # 次に bounding box で画面を切り抜く
    if candidates:
        c = candidates[0]
        clip = {
            "x": max(0, float(c["x"])),
            "y": max(0, float(c["y"])),
            "width": float(c["width"]),
            "height": float(c["height"]),
        }
        page.screenshot(path=str(out_path), clip=clip, omit_background=True)
        page.close()
        return True, f'clip:{c["selectorHint"]}'

    page.close()
    return False, "visual_not_found"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Excel準拠でポケモンずかんの画像を実用レベルで取得します。"
    )
    parser.add_argument("xlsx", help="入力Excel")
    parser.add_argument("--output-dir", default="pokemon_zukan_download_output_v3", help="出力先")
    parser.add_argument("--report-name", default="download_report.csv", help="レポートCSV名")
    parser.add_argument("--max-suffix", type=int, default=DEFAULT_MAX_SUFFIX, help="detail/xxxx-N の最大 N")
    parser.add_argument("--headless", action="store_true", help="ヘッドレスで実行")
    parser.add_argument("--skip-existing", action="store_true", help="既存画像を再取得しない")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"入力ファイルが見つかりません: {xlsx_path}", file=sys.stderr)
        return 1

    out_dir = Path(args.output_dir)
    images_root = out_dir
    report_path = out_dir / args.report_name
    out_dir.mkdir(parents=True, exist_ok=True)

    rows = extract_excel_rows(xlsx_path)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless)
        context = browser.new_context(
            viewport={"width": 1600, "height": 2200},
            device_scale_factor=2,
            locale="ja-JP",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )

        cache: dict[str, list[CandidatePage]] = {}

        with report_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([
                "row_num", "id", "answer", "selected_detail_id", "selected_detail_url",
                "selected_title", "status", "note_or_source"
            ])

            total = len(rows)
            for i, row in enumerate(rows, start=1):
                target = images_root / row.image_path
                target.parent.mkdir(parents=True, exist_ok=True)

                print(f"[{i}/{total}] {row.row_id} {row.answer}")

                if args.skip_existing and target.exists():
                    writer.writerow([row.row_num, row.row_id, row.answer, "", "", "", "skipped_existing", ""])
                    print("  -> skipped_existing")
                    continue

                try:
                    if row.base_no not in cache:
                        cache[row.base_no] = collect_candidate_pages(context, row, args.max_suffix)

                    selected, scored = choose_candidate(row, cache[row.base_no])
                    if selected is None:
                        top = " | ".join([f"{c.detail_id}:{c.title}:{score}" for score, c in scored[:5]])
                        writer.writerow([row.row_num, row.row_id, row.answer, "", "", "", "unresolved", top])
                        print("  -> unresolved")
                        continue

                    ok, source = save_main_visual(context, selected.url, target)
                    if not ok:
                        writer.writerow([
                            row.row_num, row.row_id, row.answer,
                            selected.detail_id, selected.url, selected.title, "visual_not_found", source
                        ])
                        print("  -> visual_not_found")
                        continue

                    writer.writerow([
                        row.row_num, row.row_id, row.answer,
                        selected.detail_id, selected.url, selected.title, "downloaded", source
                    ])
                    print(f"  -> saved: {target}")

                    time.sleep(0.15)

                except Exception as e:
                    writer.writerow([row.row_num, row.row_id, row.answer, "", "", "", "error", repr(e)])
                    print(f"  -> error: {e}")

        browser.close()

    print(f"完了: {report_path}")
    print(f"画像保存先: {images_root}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

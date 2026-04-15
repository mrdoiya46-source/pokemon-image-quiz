import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def normalize_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False

    text = str(value).strip().lower()
    return text in {"true", "1", "yes", "y", "on"}


def split_multi_value(value):
    if value is None:
        return []

    text = str(value).strip()
    if not text:
        return []

    for sep in ["｜", "|"]:
        text = text.replace(sep, "|")

    return [item.strip() for item in text.split("|") if item.strip()]


def normalize_string(value):
    if value is None:
        return ""
    return str(value).strip()


def row_to_question(row_dict):
    return {
        "id": normalize_string(row_dict.get("id")),
        "image": normalize_string(row_dict.get("image")),
        "answer": normalize_string(row_dict.get("answer")),
        "aliases": split_multi_value(row_dict.get("aliases")),
        "enabled": normalize_bool(row_dict.get("enabled")),
        "region": normalize_string(row_dict.get("region")).lower(),
        "tags": split_multi_value(row_dict.get("tags")),
    }


def validate_question(question, row_number):
    errors = []

    required_fields = ["id", "image", "answer", "region"]
    for field in required_fields:
      if not question.get(field):
          errors.append(f"{row_number}行目: {field} が空です")

    if not question["image"].startswith("images/"):
        errors.append(f"{row_number}行目: image は images/ から始めてください")

    return errors


def main():
    if len(sys.argv) < 2:
        print("使い方: python convert_excel_to_json.py pokemon_master_test.xlsx")
        sys.exit(1)

    input_path = Path(sys.argv[1]).resolve()
    if not input_path.exists():
        print(f"ファイルが見つかりません: {input_path}")
        sys.exit(1)

    output_path = input_path.with_name("questions.json")

    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("Excelにデータがありません。")
        sys.exit(1)

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    required_headers = {"id", "image", "answer", "aliases", "enabled", "region", "tags"}
    missing_headers = required_headers - set(headers)
    if missing_headers:
        print(f"不足している列があります: {', '.join(sorted(missing_headers))}")
        sys.exit(1)

    questions = []
    errors = []
    seen_ids = set()

    for index, row in enumerate(rows[1:], start=2):
        row_dict = dict(zip(headers, row))
        question = row_to_question(row_dict)

        row_errors = validate_question(question, index)
        errors.extend(row_errors)

        if question["id"] in seen_ids:
            errors.append(f"{index}行目: id が重複しています: {question['id']}")
        else:
            seen_ids.add(question["id"])

        questions.append(question)

    if errors:
        print("検証エラーがあります。修正してから再実行してください。")
        for error in errors:
            print(f"- {error}")
        sys.exit(1)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(questions, f, ensure_ascii=False, indent=2)

    print(f"questions.json を出力しました: {output_path}")
    print(f"問題数: {len(questions)}")


if __name__ == "__main__":
    main()